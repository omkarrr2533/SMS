import openpyxl
from io import BytesIO
from django.http import HttpResponse
from datetime import datetime


def parse_excel_headers(file):
    """
    Parse just the headers and first few rows of an Excel file.
    Returns (headers, preview_rows, error).
    """
    try:
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active

        headers = []
        preview_rows = []
        header_row_idx = None

        # Find header row (first non-empty row)
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if any(row):
                headers = [str(cell).strip() if cell else '' for cell in row]
                header_row_idx = row_idx
                break

        if not header_row_idx:
            wb.close()
            return [], [], 'Excel file is empty or has no headers'

        # Get preview rows (first 5 data rows)
        count = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx <= header_row_idx:
                continue
            if not any(row):
                continue
            preview_row = [str(cell).strip() if cell else '' for cell in row]
            preview_rows.append(preview_row)
            count += 1
            if count >= 5:
                break

        wb.close()
        return headers, preview_rows, None

    except Exception as e:
        return [], [], f'Error reading Excel file: {str(e)}'


def parse_excel_with_mapping(file, mapping):
    """
    Parse Excel file using a column mapping dict.
    mapping: {"Excel Header": "core_field_name" or "extra:key" or "skip"}

    Returns (students_data, errors, total_rows).
    Each student_data is a dict with 'core' and 'extra' keys.
    """
    try:
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active

        students = []
        errors = []
        headers = []
        header_row_idx = None

        # Find header row
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if any(row):
                headers = [str(cell).strip() if cell else '' for cell in row]
                header_row_idx = row_idx
                break

        if not header_row_idx:
            wb.close()
            return [], ['Excel file is empty or has no headers'], 0

        total_rows = 0

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx <= header_row_idx:
                continue
            if not any(row):
                continue

            total_rows += 1
            core = {}
            extra = {}
            row_errors = []

            for col_idx, value in enumerate(row):
                if col_idx >= len(headers):
                    continue
                header = headers[col_idx]
                target = mapping.get(header, 'skip')
                cell_value = str(value).strip() if value else ''

                if target == 'skip' or not target:
                    continue
                elif target.startswith('extra:'):
                    extra_key = target[6:]
                    extra[extra_key] = cell_value
                elif target == 'extra':
                    # Store with original header name as key
                    extra[header.lower().replace(' ', '_')] = cell_value
                else:
                    core[target] = cell_value

            # Validate required fields
            if not core.get('roll_no'):
                row_errors.append('Missing roll number')
            if not core.get('full_name'):
                row_errors.append('Missing student name')

            if row_errors:
                errors.append({
                    'row': row_idx,
                    'errors': row_errors,
                    'data': core,
                })
                continue

            # Generate PRN if missing
            if not core.get('prn'):
                core['prn'] = f'PRN-{core["roll_no"]}'

            students.append({
                'core': core,
                'extra': extra,
                'row': row_idx,
            })

        wb.close()
        return students, errors, total_rows

    except Exception as e:
        return [], [{'row': 0, 'errors': [str(e)], 'data': {}}], 0


def parse_excel(file):
    """
    Legacy parse function for backward compatibility.
    Uses auto-suggest mapping from field_registry.
    """
    try:
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active

        students = []
        errors = []
        headers = []
        header_row_idx = None

        # Find header row (first non-empty row)
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if any(row):
                headers = [str(cell).strip().lower() if cell else '' for cell in row]
                header_row_idx = row_idx
                break

        if not header_row_idx:
            return [], ['Excel file is empty or has no headers']

        # Process data rows
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx <= header_row_idx:
                continue
            if not any(row):
                continue

            # Map row data to dict using headers
            row_data = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers):
                    row_data[headers[col_idx]] = str(value).strip() if value else ''

            # Extract student data - flexible header matching
            student = {
                'roll_no': row_data.get('roll_no', row_data.get('roll no', row_data.get('rollno', row_data.get('roll', row_data.get('student roll no', ''))))),
                'prn': row_data.get('prn', row_data.get('prn no', row_data.get('prn_no', row_data.get('prn no.', '')))),
                'abc_id': row_data.get('abc_id', row_data.get('abc id', '')),
                'full_name': row_data.get('full_name', row_data.get('name', row_data.get('full name', row_data.get('student name', row_data.get('student full name', ''))))),
                'phone': row_data.get('phone', row_data.get('phone no', row_data.get('mobile', row_data.get('phone_no', row_data.get('contact number (student)', ''))))),
                'email': row_data.get('email', row_data.get('email id', row_data.get('email_id', ''))),
                'parent_name': row_data.get('parent_name', row_data.get('parent name', row_data.get('guardian name', ''))),
                'parent_phone': row_data.get('parent_phone', row_data.get('parent phone', row_data.get('parent mobile', row_data.get('parent contact number (father)', '')))),
                'birthdate': row_data.get('birthdate', row_data.get('birth date', row_data.get('dob', ''))),
                'gender': row_data.get('gender', '').strip().lower() if row_data.get('gender', '').strip() else '',
                'address': row_data.get('address', row_data.get('local address', '')),
                'permanent_address': row_data.get('permanent_address', row_data.get('permanent address', '')),
            }

            # Validate required fields
            if not student['roll_no']:
                errors.append(f'Row {row_idx}: Missing roll number')
                continue
            if not student['full_name']:
                errors.append(f'Row {row_idx}: Missing student name')
                continue

            # Generate PRN if missing
            if not student['prn']:
                student['prn'] = f'PRN-{student["roll_no"]}'

            students.append(student)

        wb.close()
        return students, errors

    except Exception as e:
        return [], [f'Error reading Excel file: {str(e)}']


def export_students_excel(students, file_name='students'):
    """Export student list as Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    # Headers
    headers = ['Roll No', 'PRN', 'ABC ID', 'Full Name', 'Phone', 'Email',
               'Parent Name', 'Parent Phone', 'Birthdate', 'Gender',
               'Local Address', 'Permanent Address', 'Class', 'Division',
               'Year', 'Status']

    # Collect extra data headers from all students
    extra_keys = []
    seen_keys = set()
    for student in students:
        if hasattr(student, 'extra_data') and student.extra_data:
            for key in student.extra_data.keys():
                if key not in seen_keys:
                    extra_keys.append(key)
                    seen_keys.add(key)

    headers.extend([k.replace('_', ' ').title() for k in extra_keys])
    ws.append(headers)

    # Bold headers
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Data
    for student in students:
        row = [
            student.roll_no,
            student.prn,
            student.abc_id,
            student.full_name,
            student.phone,
            student.email,
            student.parent_name,
            student.parent_phone,
            student.birthdate.strftime('%Y-%m-%d') if student.birthdate else '',
            student.get_gender_display(),
            student.address,
            student.permanent_address,
            student.class_name,
            student.division,
            student.year,
            student.get_status_display(),
        ]
        # Append extra data columns
        extra = student.extra_data if hasattr(student, 'extra_data') and student.extra_data else {}
        for key in extra_keys:
            row.append(extra.get(key, ''))

        ws.append(row)

    # Auto-width columns
    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'
    wb.save(response)
    return response


def export_dynamic_excel(students, columns, file_name='students'):
    """Export with only selected columns. columns is a list of field names (or extra:key)."""
    from .field_registry import CORE_FIELDS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    # Build headers
    headers = []
    for col in columns:
        if col.startswith('extra:'):
            headers.append(col[6:].replace('_', ' ').title())
        elif col in CORE_FIELDS:
            headers.append(CORE_FIELDS[col]['label'])
        else:
            headers.append(col.replace('_', ' ').title())
    ws.append(headers)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Data
    for student in students:
        row = []
        for col in columns:
            if col.startswith('extra:'):
                key = col[6:]
                extra = student.extra_data if hasattr(student, 'extra_data') and student.extra_data else {}
                row.append(extra.get(key, ''))
            elif col == 'gender':
                row.append(student.get_gender_display())
            elif col == 'birthdate':
                row.append(student.birthdate.strftime('%Y-%m-%d') if student.birthdate else '')
            elif col == 'status':
                row.append(student.get_status_display())
            else:
                row.append(getattr(student, col, ''))
        ws.append(row)

    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'
    wb.save(response)
    return response


def export_dynamic_pdf(students, columns, file_name='students'):
    """Export selected columns as PDF."""
    from .field_registry import CORE_FIELDS
    from django.template.loader import render_to_string

    headers = []
    for col in columns:
        if col.startswith('extra:'):
            headers.append(col[6:].replace('_', ' ').title())
        elif col in CORE_FIELDS:
            headers.append(CORE_FIELDS[col]['label'])
        else:
            headers.append(col.replace('_', ' ').title())

    rows = []
    for student in students:
        row = []
        for col in columns:
            if col.startswith('extra:'):
                key = col[6:]
                extra = student.extra_data if hasattr(student, 'extra_data') and student.extra_data else {}
                row.append(extra.get(key, ''))
            elif col == 'gender':
                row.append(student.get_gender_display())
            elif col == 'birthdate':
                row.append(student.birthdate.strftime('%Y-%m-%d') if student.birthdate else '')
            elif col == 'status':
                row.append(student.get_status_display())
            else:
                row.append(getattr(student, col, ''))
        rows.append(row)

    try:
        from xhtml2pdf import pisa

        html_string = render_to_string('students/export_pdf.html', {
            'headers': headers,
            'rows': rows,
            'file_name': file_name,
        })

        result = BytesIO()
        pdf = pisa.CreatePDF(BytesIO(html_string.encode('utf-8')), dest=result)

        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{file_name}.pdf"'
            return response

    except ImportError:
        pass

    # Fallback to Excel if PDF generation fails
    return export_dynamic_excel(students, columns, file_name)
